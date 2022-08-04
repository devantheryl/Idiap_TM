#!/usr/bin/env python3

# Filename: pycalc.py

# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""PyCalc is a simple calculator built using Python and PyQt5."""


import pandas as pd
import numpy as np
from openpyxl import load_workbook
import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import get_column_letter
import csv
import re
from pandas.tseries.offsets import DateOffset
import json
import os
from datetime import datetime, timedelta, date
from itertools import islice

import time
import json
from collections import deque

from src.TF_env3 import TF_environment
from tensorforce import Environment, Runner, Agent

import src.utils as utils

import xlwings as xw


import sys

from functools import partial

# Import QApplication and the required widgets from PyQt5.QtWidgets
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication
from PyQt5.QtWidgets import QGridLayout
from PyQt5.QtWidgets import QLineEdit
from PyQt5.QtWidgets import QMainWindow
from PyQt5.QtWidgets import QPushButton
from PyQt5.QtWidgets import QVBoxLayout
from PyQt5.QtWidgets import QHBoxLayout
from PyQt5.QtWidgets import QWidget

__version__ = "0.1"
__author__ = "Leodanis Pozo Ramos"

ERROR_MSG = "ERROR"


# Create a subclass of QMainWindow to setup the calculator's GUI
class PyCalcUi(QMainWindow):
    """PyCalc's View (GUI)."""

    def __init__(self):
        """View initializer."""
        super().__init__()
        # Set some main window's properties
        self.setWindowTitle("PyCalc")
        self.setFixedSize(235, 235)
        # Set the central widget and the general layout
        self.generalLayout = QVBoxLayout()
        self._centralWidget = QWidget(self)
        self.setCentralWidget(self._centralWidget)
        self._centralWidget.setLayout(self.generalLayout)
        # Create the display and the buttons
        self._createDisplay()
        self._createButtons()

    def _createDisplay(self):
        """Create the display."""
        # Create the display widget
        self.display_layout = QHBoxLayout()
        self.batch = QLineEdit()
        self.formulation = QLineEdit()
        self.scale = QLineEdit()
        self.target = QLineEdit()
   
        # Add the display to the general layout
        self.display_layout.addWidget(self.batch)
        self.display_layout.addWidget(self.formulation)
        self.display_layout.addWidget(self.scale)
        self.display_layout.addWidget(self.target)
        
        self.generalLayout.addLayout(self.display_layout)

    def _createButtons(self):
        """Create the buttons."""
        self.buttons = {}
        buttonsLayout = QGridLayout()
        # Button text | position on the QGridLayout
        buttons = {
            "Run": (0, 0),
            "Reset": (0, 1),
            
        }
        # Create the buttons and add them to the grid layout
        for btnText, pos in buttons.items():
            self.buttons[btnText] = QPushButton(btnText)
            self.buttons[btnText].setFixedSize(40, 40)
            buttonsLayout.addWidget(self.buttons[btnText], pos[0], pos[1])
        # Add buttonsLayout to the general layout
        self.generalLayout.addLayout(buttonsLayout)

    def setDisplayText(self, text):
        """Set display's text."""
        self.display.setText(text)
        self.display.setFocus()

    def displayText(self):
        """Get display's text."""
        return self.display.text()

    def clearDisplay(self):
        """Clear the display."""
        self.setDisplayText("")


# Create a Model to handle the calculator's operation
def evaluateExpression(expression):
    """Evaluate an expression."""
    try:
        result = str(eval(expression, {}, {}))
    except Exception:
        result = ERROR_MSG

    return result


# Create a Controller class to connect the GUI and the model
class PyCalcCtrl:
    """PyCalc's Controller."""

    def __init__(self, model, view):
        """Controller initializer."""
        self._evaluate = model
        self._view = view
        # Connect signals and slots
        self._connectSignals()

    def _runPlanning(self):
        """Evaluate expressions."""
        
        print(self._view.batch.text())
        print("working")
        #load the plannign
        wb = load_workbook(filename = "C:/Users/LDE/Prog/projet_master/digital_twins/data/Planning_Production_2022.xlsm",data_only = True, keep_vba = True)
        ws = wb["PLANNING"]
        
        #unmerge the cell in the file
        mergedcells =[]  
        for group in ws.merged_cells.ranges:
            mergedcells.append(group)
        
        for group in mergedcells:
            min_col, min_row, max_col, max_row = group.bounds 
            top_left_cell_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(group))   # you need to unmerge before writing (see explanation #1 below)
            for irow in range(min_row, max_row+1):
                for jcol in range(min_col, max_col+1): 
                    ws.cell(row = irow, column = jcol, value = top_left_cell_value)
                    
        #now "ws" is the excel file unmerged
        #convert openpyxl format to pandas
        data = ws.values
        cols = next(data)[1:]
        data = list(data)
        idx = [r[0] for r in data]
        data = (islice(r, 1, None) for r in data)
        df = pd.DataFrame(data, index=idx, columns=cols)
        
        #keep a trace of the original plannification dataframe
        df_original = df.copy()
        
        
        #keep only machines and operator index
        df = df[~df.index.duplicated(keep='first')]
        index_to_keep = df.index.dropna()
        
        df = df.loc[index_to_keep]
        
        df =df.drop(df.filter(like='None').columns, axis=1)
        df = df.drop(columns = df.columns[0])
        
        #resample the column of df
        columns = pd.to_datetime(df.columns)
        
        resampled_date = pd.DataFrame(0,columns=[""],index=columns).resample('12H', closed = "left").mean().index.tolist()
        df = df.iloc[: , :-1] #drop the last column to make the datetime list fit
        df.columns = resampled_date
        df_t = df.transpose() #transpose to have the date as index
        df_t = df_t.drop(columns = df_t.columns[0])#again drop useless column
        df_t = df_t.fillna('0')
        
        
        #extract_machine
        machines = ['Broyage polymère B1', 'Broyage polymère B2', 'Tamisage polymère B2',
               'Mélanges B1 ', 'Extrusion B1', 'Mélanges B2', 'Extrusion B2',
               'Broyage bâtonnets B1 ', 'Broyage bâtonnets B2 ',
               'Tamisage Microgranules B2', 'Combin. des fractions de microgranules',
               'Milieu de suspension  ', 'Remplissage Poudre + liquide B2']
        machine_dict = {"m1" : ['Broyage polymère B1','Broyage bâtonnets B1 ' ],
                        "m2" : ['Broyage polymère B2','Broyage bâtonnets B2 '],
                        "m3" : ['Tamisage polymère B2', 'Tamisage Microgranules B2'],
                        "m4" : ['Mélanges B1 ', 'Mélanges B2'],
                        "m5" : ['Extrusion B1','Extrusion B2'],
                        "m6" : ['Combin. des fractions de microgranules'],
                        "m7" : ['Milieu de suspension  '],
                        "m8" : ['Remplissage Poudre + liquide B2']}
        
        operator_machine_dict = {"m1" : 2,
                                 "m2" : 2,
                                 "m3" : 2,
                                 "m4" : 3,
                                 "m5" : 2,
                                 "m6" : 2,
                                 "m7" : 2,
                                 "m8" : 8}
        
        df_machine = df_t[machines].copy()
        
        #merge same step using the same machine
        column_to_use = []
        df_machine["operator"] = 0
        column_to_use.append("operator")
        for key, value in machine_dict.items():
            df_machine[key] = "0"
            column_to_use.append(key)
            for machine in value:
                #convert text to string "1"
                df_machine.loc[df_machine[machine] != "0", machine] = "1"
                #merge the machines
                df_machine.loc[df_machine[machine] == "1",key] = "1"
                
                df_machine[machine] = df_machine[machine].astype(int)
            df_machine[key] = df_machine[key].astype(int)
            df_machine["operator"] = df_machine["operator"] + df_machine[key] * operator_machine_dict[key]
                
        df_machine_final = df_machine[column_to_use]
        
        target = self._view.target.text()
        target = pd.to_datetime(target)
        futur_length = 100
        end = target + DateOffset(hours = 12*4)#4 = perry duration
        start = end - DateOffset(hours = 12*(futur_length-1))
        
        df_futur_machine = df_machine_final.loc[start:end].copy()
        index_to_keep = df_futur_machine.index + DateOffset(hours = 11, minutes = 59)
        df_futur_machine.set_index(index_to_keep,inplace = True)
        
        df_futur_machine = df_futur_machine.iloc[::-1]
        df_futur_machine.drop(df_futur_machine.index[0], inplace = True)
        
        df_futur_machine["days"] = df_futur_machine.index.dayofweek
        df_futur_machine.operator = np.where(df_futur_machine.days < 5, 12 - df_futur_machine.operator, 0)
        
        df_futur_machine = df_futur_machine[column_to_use]
        
        
        directory = "model/reccurent_job_ddqn_weekend_final/pleasant-voice-9"
        filename = "0000489888.hdf5"
        
        #model params
        operator_vector_length = 7
        echu_weights = -97
        ordo_weights = 2
        job_finished_weigths = 30
        forward_weights = 0
        independent = True
        
        
        targets = [target]
        formulations = [int(self._view.formulation.text())]
        echelles = [int(self._view.scale.text())]
        job_names = [self._view.batch.text()]
        
        
        target = targets[0]
        formulation = formulations[0]
        echelle = echelles[0]
        job_name = job_names[0]
        nbr_operation_max = 15
        nbr_machines = 8
        nbr_operator = 12
        nbr_job_to_use = len(targets)
        
        
        environment = Environment.create(environment = TF_environment(target, formulation,echelle, job_name, nbr_operation_max, nbr_machines, nbr_operator, 
                                                                          operator_vector_length,None, echu_weights = echu_weights,
                                                                          forward_weights = forward_weights, ordo_weights = ordo_weights,
                                                                          job_finished_weigths = job_finished_weigths, independent =False))
            
        
        agent = Agent.load(
                directory = directory,
                filename = filename,
                environment = environment,
                )
        
        
        planning_tot_max = None
        lead_time_tot = []
        for j in range(nbr_job_to_use):
            lead_time_min = 1000
            print("job : ", j)
            for tentative in range(20):
                reward_tot = 0
                
                futur_state = df_futur_machine
                planning_tot = None
                
                echu = False
        
                # Initialize episode
                target = pd.to_datetime(targets[j])
                formulation = formulations[j]
                echelle = echelles[j]
                job_name = job_names[j]
        
                environment.job_name = job_name
                environment.target = target
                environment.formulation = formulation 
                environment.echelle = echelle
                environment.futur_state = futur_state
                states = environment.reset()
                reward_batch = 0
                terminal = False
                internals = agent.initial_internals()
        
                while not terminal:
                     # Episode timestep
                    actions, internals = agent.act(states=states, internals = internals, independent=True, deterministic = False)
                    states, terminal, reward = environment.execute(actions=actions)
        
                    reward_batch += reward
        
                futur_state = environment.get_env().state_full
                reward_tot += reward_batch
                planning = environment.get_env().get_gant_formated()
                
                count = (planning['Start'] == 0).sum()
                if count == 0:
                    lead_time = environment.get_env().job.lead_time
                    print("tentative : ", tentative, "  : no break", " lead_time : ", lead_time)
                    if lead_time < lead_time_min:
                        print("new plann : ", np.mean(lead_time))
        
                        planning_max = planning.copy()
                        futur_state_max = futur_state.copy()
                        lead_time_min = lead_time
            planning_tot_max  = pd.concat([planning_tot_max,planning_max])
            df_futur_machine = futur_state_max
            lead_time_tot.append(lead_time_min)
            
        print("lead time mean : ",np.mean(lead_time_tot))
            
        
        #utils.visualize(planning_tot,historic_time_tot,historic_operator_tot)
        
        
        operation_machine = {
                         (0) : "echu",
                         (1,1) : "Broyage polymère B1",
                         (1,2) : "Broyage polymère B2",
                         (2,1) : "Broyage polymère B1",
                         (2,2) : "Broyage polymère B2",
                         (3,3) : "Tamisage polymère B2",
                         (4,3) : "Tamisage polymère B2",
                         (5,4) : "Mélanges B1 ",
                         (6,4) : "Mélanges B1 ",
                         (7,5) : "Extrusion B2",
                         (8,5) : "Extrusion B2",
                         (9,1) : "Broyage bâtonnets B1 ",
                         (9,2) : "Broyage bâtonnets B2 ",
                         (10,1) : "Broyage bâtonnets B1 ",
                         (10,2) : "Broyage bâtonnets B2 ",
                         (11,3) : "Tamisage Microgranules B2",
                         (12,3) : "Tamisage Microgranules B2",
                         (13,7) : "Milieu de suspension  ",
                         (14,6) : "Combin. des fractions de microgranules",
                         (15,8) : "Remplissage Poudre + liquide B2",
                         }
        
        planning_tot_reformated = planning_tot_max.copy()
        planning_tot_reformated.Finish = planning_tot_reformated.Finish - DateOffset(hours = 11, minutes = 59)
        planning_tot_reformated.Start = planning_tot_reformated.Start + DateOffset(minutes = 1)
        
        #link operation and machine
        op_machine = []
        for index, row in planning_tot_reformated.iterrows():
            
            op = row["Operation"]
            machine = row["Machine"]
            duration = row["Duration"]
            if duration != 0:
                op_machine.append(operation_machine[int(op),int(machine)])
            else:
                op_machine.append(operation_machine[0])
        
        planning_tot_reformated["op_machine"] = op_machine
        
        start = planning_tot_reformated.Start.iloc[0]
        start_min = planning_tot_reformated.Start.min()
        fin = planning_tot_reformated.Finish.iloc[0] 
        fin_max = planning_tot_reformated.Finish.max()
        
        planning_tot_reformated.sort_values("Start",inplace = True, ignore_index = True)
        print(planning_tot_reformated)
        
        print(start_min, fin_max)
        
        excel_out = df_t.copy()
        
        for index, row in planning_tot_reformated.iterrows():
            op = row["op_machine"]
            start = row["Start"]
            end = row["Finish"]
            
            excel_out.loc[start:end, op] = "FINAL -TEST BATCH"
            
        
        excel_out.to_excel("final_planning.xlsx")
        
        
        operation_machine_excel_rows = {

                 "Broyage polymère B1": 4,
                 "Broyage polymère B2" : 13,
                 "Tamisage polymère B2" : 22,          
                 "Mélanges B1 " : 29,
                 "Extrusion B2" : 54,
                 "Broyage bâtonnets B1 " : 61,
                 "Broyage bâtonnets B2 " : 71,
                 "Tamisage Microgranules B2" : 81,
                 "Milieu de suspension  " : 98,
                 "Combin. des fractions de microgranules" : 88,
                 "Remplissage Poudre + liquide B2" : 104,
                 }

        wb = xw.Book("C:/Users/LDE/Prog/projet_master/digital_twins/data/Planning_Production_2022_test.xlsm")
        ws = wb.sheets["PLANNING"]
        
        rng = ws.range("A1:AQN1")
        
        
        print(rng.shape)
        
        #create the color for each job
        color_dict = {}
        for job in job_names:
            color = tuple(np.random.choice(range(256), size=3))
            color_dict[job] = color
        
        
        
        for r in rng:
            
            current_date = pd.to_datetime(r.value)
            if current_date is not None:
                for i, row in planning_tot_reformated.iterrows():
                    afternoon = False
                    op_date = row.at["Start"]
                    if current_date.day_of_year == op_date.day_of_year:
                        op = row.at["op_machine"]
                        excel_row_index = operation_machine_excel_rows[op]
                        if op_date.hour != 0:
                            r = r.offset(column_offset = 1)
                            afternoon = True
                        
                        duration = row["Duration"].days * 2 + row["Duration"].seconds//3600 //12
                        print(duration)
                        last_excel_cell = 0
                        for j in range(duration):
                            if afternoon:
                                if j > 0:
                                    colum_letter = get_column_letter(r.offset(column_offset = j+1).column)
                                else:
                                    colum_letter = get_column_letter(r.offset(column_offset = j).column)
                            else:
                                if j > 1:
                                    colum_letter = get_column_letter(r.offset(column_offset = j+1).column)
                                    
                                else:
                                    colum_letter = get_column_letter(r.offset(column_offset = j).column)
                                    
                            excel_cell = colum_letter + str(excel_row_index)
                            
                            
                            
                            if (duration == 2 or duration ==4) and (j == 1 or j == 3):
                                ws.range(last_excel_cell + ":" + excel_cell).merge()
                            else:
                                ws.range(excel_cell).value = row["Job"]
                                ws.range(excel_cell).color = color_dict[row["Job"]]
                                
                                
                            last_excel_cell = excel_cell
                    
                



    def _connectSignals(self):
        """Connect signals and slots."""

        self._view.buttons["Run"].clicked.connect(partial(self._runPlanning))
        

        


# Client code
def main():
    """Main function."""
    # Create an instance of `QApplication`
    pycalc = QApplication(sys.argv)
    # Show the calculator's GUI
    view = PyCalcUi()
    view.show()
    # Create instances of the model and the controller
    model = evaluateExpression
    PyCalcCtrl(model=model, view=view)
    # Execute calculator's main loop
    sys.exit(pycalc.exec_())


if __name__ == "__main__":
    main()
