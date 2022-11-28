# -*- coding: utf-8 -*-
"""
Created on Fri Dec 17 16:45:59 2021

@author: LDE
"""

from datetime import datetime, timedelta

import numpy as np
import xlwings as xw
from openpyxl import load_workbook

from openpyxl.utils.cell import get_column_letter
from itertools import islice
import pandas as pd
from unidecode import unidecode
from pandas.tseries.offsets import DateOffset

def visualize(excel_file, batch_names, planning_tot_final, selected_operators, planning_octo_laverie):
    
    operation_machine_excel_rows = {
    
                 "Broyage polymère B1": 4,
                 "Broyage polymère B2" : 13,
                 "Tamisage polymère B2" : 22,          
                 "Mélanges B1 " : 29,
                 "Mélanges B2" : 45,
                 "Extrusion B2" : 54,
                 "Broyage bâtonnets B1 " : 61,
                 "Broyage bâtonnets B2 " : 71,
                 "Tamisage Microgranules B2" : 81,
                 "Milieu de suspension  " : 98,
                 "Combin. des fractions de microgranules" : 88,
                 "Remplissage Poudre + liquide B2" : 104,
                 "Capsulage" : 121,
                 "Inspection visuelle " :129 
                 }
        
    operator_excel_rows = {
        "SFR" : 192, 
        "BPI" : 194, 
        "JPI" : 196, 
        "JDD" : 198, 
        "FDS" : 200, 
        "SFH" : 202, 
        "NDE" : 204, 
        "LTF" : 206, 
        "SRG" : 208, 
        "JPE" : 210, 
        "RPI" : 212, 
        "ANA" : 214,
        "MTR" : 216, 
        "CMT" : 218, 
        "CGR" : 220, 
        "CPO" : 222, 
        "REA" : 224,
        "NRO" : 227,
        "VZU" : 229,
        "ACH" : 231,
        "LBU" : 233
    }
    
    operation_occupations = {
        "Broyage polymère B1": "BP",
        "Broyage polymère B2" : "BP",
        "Tamisage polymère B2" : "TP",          
        "Mélanges B1 " : "MEL",
        "Mélanges B2" : "MEL",
        "Extrusion B2" : "EXT",
        "Broyage bâtonnets B1 " : "BB",
        "Broyage bâtonnets B2 " : "BB",
        "Tamisage Microgranules B2" : "TM",
        "Milieu de suspension  " : "MILIEU",
        "Combin. des fractions de microgranules" : "CF",
        "Remplissage Poudre + liquide B2" : "LYO",
        "Capsulage" : "CAPS",
        "Inspection visuelle " : "IV"
    }


    wb = xw.Book(excel_file)
    ws = wb.sheets["PLANNING"]
    
    rng = ws.range("A1:AQN1")
    
    
    #create the color for each job
    color_dict = {}
    for job in np.sort(np.unique(planning_tot_final["Job"].tolist())):
        color = tuple(np.random.choice(range(256), size=3))
        if "75:25" in job:
            color_dict[job] = color_dict[job.split(" ")[0]]
        else:
            color_dict[job] = color
    
    
    morning = True
    for r in rng:
        
        current_date = pd.to_datetime(r.value)
        
        if current_date is None:
            morning = True
        else:
            if morning:
                try:
                    operators = selected_operators.loc[current_date:current_date + DateOffset(hours = 12)].T
                    operators = operators.rename(index=operator_excel_rows)
                    
                    column_letter_morning = get_column_letter(r.column)
                    column_letter_afternoon = get_column_letter(r.offset(column_offset = 1).column)
                    
                    excel_cell_operators_morning = [column_letter_morning + str(index_operator) for index_operator in operators.index.to_list()]
                    excel_cell_operators_afternoon = [column_letter_afternoon + str(index_operator) for index_operator in operators.index.to_list()]
                    
                    occ_morning = operators.T.iloc[0].to_list()
                    occ_afternoon = operators.T.iloc[1].to_list()
                    
                    for i in range(len(excel_cell_operators_morning)):
                        if occ_morning[i] == "0":
                            occ_morning[i] = ""
                        if occ_afternoon[i] == "0":
                            occ_afternoon[i] = ""
                        ws.range(excel_cell_operators_morning[i]).value = occ_morning[i]
                        ws.range(excel_cell_operators_afternoon[i]).value = occ_afternoon[i]
                except:
                    pass
                    
            
            for i, row in planning_tot_final.iterrows():
                afternoon = False
                op_date = row.at["Start"]
                if current_date.day_of_year == op_date.day_of_year:
                    op = row.at["op_machine"]
                    excel_row_index_operation = operation_machine_excel_rows[op]
                    
                    if op_date.hour != 0:
                        r_aft = r.offset(column_offset = 1)
                        afternoon = True
                    
                    duration = row["Duration"].days * 2 + row["Duration"].seconds//3600 //12
                    last_excel_cell_operation = 0
                    for j in range(duration):
                        if afternoon:
                            if j > 0:
                                colum_letter = get_column_letter(r_aft.offset(column_offset = j+1).column)
                            else:
                                colum_letter = get_column_letter(r_aft.offset(column_offset = j).column)
                        else:
                            if j > 1:
                                colum_letter = get_column_letter(r.offset(column_offset = j+1).column)
                                
                            else:
                                colum_letter = get_column_letter(r.offset(column_offset = j).column)
                                
                        excel_cell_operation = colum_letter + str(excel_row_index_operation)
                        
                        if (duration == 2 or duration ==4) and (j == 1 or j == 3):
                            ws.range(last_excel_cell_operation + ":" + excel_cell_operation).merge()
                                
                        else:
                            
                            ws.range(excel_cell_operation).value = row["Job"]
                            ws.range(excel_cell_operation).color = color_dict[row["Job"]]

                            
                        last_excel_cell_operation = excel_cell_operation
                        
                        
            morning = False
        
    
    

def get_new_time(current_time, delta):
    temp = current_time
    
    for i in range(abs(delta)):
        if delta < 0:
            if temp.time().hour == 12:
                temp -= timedelta(hours = 12)
            else:
                temp -= timedelta(days=1)
                temp = temp.replace(hour = 12)
        else :
            if temp.time().hour == 00:
                temp += timedelta(hours = 12)
            else:
                temp += timedelta(days=1)
                temp = temp.replace(hour = 00)
            pass
    return temp


def get_delta_time(current_time, target_time):
    
    
    if current_time > target_time:
        temp = current_time
        i = 0
        while(1):
            if temp == target_time:
                return i
            temp = get_new_time(temp,-1)
            i +=1
    
    if current_time < target_time:
        i = 0
        temp = target_time
        while(1):
            if temp == current_time:
                return -i
            temp = get_new_time(temp,-1)
            i+=1
        
    return 0
    

def generate_test_scenarios(start_date, nbr_job, seed):
    day_stat = [0.25,0.25,0.25,0.25]
    formu_stat = [0.25,0.25,0.5]
    
    target_dates = {}
    start_date = datetime.fromisoformat(start_date)
    np.random.seed(seed)
    i=0
    while i < nbr_job:
        day = np.random.choice([0,1,2,3], 1, p = day_stat)[0]
        
        #if jeudi
        if day == 3:
            formu = np.random.choice([1,3,6], 1, p = formu_stat)[0]
            target_dates[str(start_date)] = formu #on ajoute aussi le lundi
            if i+1<nbr_job:
                i += 1
            else:
                return target_dates
            
        formu = np.random.choice([1,3,6], 1, p = formu_stat)[0]
        target_date = get_new_time(start_date, 2*day)
        target_dates[str(target_date)] = formu
        start_date = get_new_time(start_date, 14) # + 1 week
        
        i +=1
        
    return target_dates


def unmerge_excel(excel_file):
    #load the plannign
    wb = load_workbook(filename = excel_file ,data_only = True, keep_vba = True)
    ws = wb["PLANNING"]
    
    #unmerge the cell in the file
    mergedcells =[]  
    for group in ws.merged_cells.ranges:
        mergedcells.append(group)
    
    for group in mergedcells:
        min_col, min_row, max_col, max_row = group.bounds 
        top_left_cell_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(group))   # you need to unmerge before writing (see explanation #1 below)
        for irow in range(min_row, max_row+1):
            for jcol in range(min_col, max_col+1): 
                ws.cell(row = irow, column = jcol, value = top_left_cell_value)
                
    #now "ws" is the excel file unmerged
    #convert openpyxl format to pandas
    data = ws.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)
    df_restrict = df.copy()
    
    
    """
    prepare the full df 
    """
    df = df.reset_index()
    df = df.set_index([df.columns[0], df.columns[1]])
    

    #keep only machines and operator index
    index_to_keep = df.index.dropna()
    
    df = df.loc[index_to_keep]
    
    df =df.drop(df.filter(like='None').columns, axis=1)
    
    #resample the column of df
    columns = pd.to_datetime(df.columns)
    
    resampled_date = pd.DataFrame(0,columns=[""],index=columns).resample('12H', closed = "left").mean().index.tolist()
    df = df.iloc[: , :-1] #drop the last column to make the datetime list fit
    df.columns = resampled_date
    df_t = df.transpose() #transpose to have the date as index
    df_t = df_t.drop(columns = df_t.columns[0])#again drop useless column
    df_t = df_t.fillna('0')
    
    """
    prepare the resticted df 
    """
    #keep only machines and operator index
    df_restrict = df_restrict[~df_restrict.index.duplicated(keep='first')]
    index_to_keep = df_restrict.index.dropna()
    
    df_restrict = df_restrict.loc[index_to_keep]
    
    df_restrict =df_restrict.drop(df_restrict.filter(like='None').columns, axis=1)
    df_restrict = df_restrict.drop(columns = df_restrict.columns[0])
    
    #resample the column of df
    columns = pd.to_datetime(df_restrict.columns)
    
    resampled_date = pd.DataFrame(0,columns=[""],index=columns).resample('12H', closed = "left").mean().index.tolist()
    df_restrict = df_restrict.iloc[: , :-1] #drop the last column to make the datetime list fit
    df_restrict.columns = resampled_date
    df_restrict_t = df_restrict.transpose() #transpose to have the date as index
    df_restrict_t = df_restrict_t.drop(columns = df_restrict_t.columns[0])#again drop useless column
    df_restrict_t = df_restrict_t.fillna('0')
    
    
    return df_t, df_restrict_t

def extract_machine_operator_state(df):
    
    #extract_machine
    machines = ['Broyage polymère B1', 'Broyage polymère B2', 'Tamisage polymère B2',
           'Mélanges B1 ', 'Extrusion B1', 'Mélanges B2', 'Extrusion B2',
           'Broyage bâtonnets B1 ', 'Broyage bâtonnets B2 ',
           'Tamisage Microgranules B2', 'Combin. des fractions de microgranules',
           'Milieu de suspension  ', 'Remplissage Poudre + liquide B2', 'Sortie Lyo','Capsulage']
    
    operators = ["SFR", "BPI", "JPI", "JDD", "FDS", "SFH", "NDE", "LTF", "SRG", "JPE", "RPI", "ANA",
                "MTR", "CMT", "CGR", "CPO", "REA", "NRO", "VZU", "ACH"]
    
    
    merge_occupations = ["FORMATION","OCTODURE","LAVERIE","LAVAGE", "BP", "TP", "MEL", "EX",  "BB", "TM", "CF", "MILIEU", 
                     "LYO", "NETTOYAGE", "CAPS", "IV", "ETIQUE", 
                     "REQUALIF","TEST", "VACANCE", "ABSENT" ,"CONGE",]
    
    vacances_occupations = ["VACANCE", "ABSENT" ,"CONGE","FORMATION", "IV"]
    
    machine_dict = {"m0" : ['Broyage polymère B1','Broyage bâtonnets B1 ' ],
                    "m1" : ['Broyage polymère B2','Broyage bâtonnets B2 '],
                    "m2" : ['Tamisage polymère B2', 'Tamisage Microgranules B2'],
                    "m3" : ['Mélanges B1 ', 'Mélanges B2','Extrusion B1','Extrusion B2'],
                    "m4" : ['Mélanges B1 ', 'Mélanges B2','Extrusion B1','Extrusion B2'],
                    "m5" : ['Combin. des fractions de microgranules'],
                    "m6" : ['Milieu de suspension  '],
                    "m7" : ['Remplissage Poudre + liquide B2'],
                    "m8" : ['Sortie Lyo'],
                    "m9":  ['Capsulage']
                    
                    }
    
    vacances_conge2022 = ["2022-04-18","2022-05-26","2022-05-27", "2022-06-06","2022-06-16", "2022-08-01", "2022-08-15",
                          "2022-11-01", "2022-12-08", "2022-12-26","2022-12-27", "2022-12-28", "2022-12-29", "2022-12-30"]
    
    operator_machine_dict = {"m0" : 2,
                             "m1" : 2,
                             "m2" : 2,
                             "m3" : 3,
                             "m4" : 2,
                             "m5" : 2,
                             "m6" : 2,
                             "m7" : 8,
                             "m8" : 2,
                             "m9" : 2,
                             }
    
    columns_name = df.columns
    df_machine = df[machines].copy()
    
    #merge same step using the same machine
    column_to_use = ["m0", "m1", "m2", "m3", "m4", "m5", "m6", "m7", "m8", "m9"]
    
    
    #EXTRAIT L'UTILISATION DES MACHINES + LES OPERATEURS UTILISé PAR LES OPERATIONS PLANNIFIéES
    for key, value in machine_dict.items():
        df_machine[key] = "0"

        for machine in value:
            
            #machine utilisée si la case != '0'
            df_machine.loc[df_machine[machine] != "0", key] = machine

        
        
        
    df_ressources_final = df_machine[column_to_use].copy()
    
    
    #EXTRAIT LES TACHE EFFECTUéE PAR LES OPéRATEURS
    df_operator = df[operators].copy()
    
    #normalise les noms des occupations
    for operator in operators:
        df_operator[operator] = df_operator[operator].apply(unidecode).str.upper()
        for potential in merge_occupations:
            df_operator.loc[df_operator[operator].str.contains(potential),operator] = potential
                


    
    df_ressources_final = df_ressources_final[column_to_use]
    
    
    return df_ressources_final, df_operator




