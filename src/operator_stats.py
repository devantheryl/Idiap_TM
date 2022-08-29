# -*- coding: utf-8 -*-
"""
Created on Mon Aug 29 14:46:50 2022

@author: LDE
"""
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import get_column_letter
import csv
import re
from pandas.tseries.offsets import DateOffset
import json
import os
from datetime import datetime, timedelta, date
from itertools import islice

import unidecode

import time
import json
from collections import deque

from src.TF_env3 import TF_environment
from tensorforce import Environment, Runner, Agent

import src.utils as utils

import xlwings as xw

import matplotlib.pyplot as plt
import pandas as pd

def get_operators_stats(filename):
   
    #load the plannign
    wb = load_workbook(filename = filename, data_only = True, keep_vba = True)
    ws = wb["PLANNING"]
    
    #unmerge the cell in the file
    mergedcells =[]  
    for group in ws.merged_cells.ranges:
        mergedcells.append(group)
    
    for group in mergedcells:
        min_col, min_row, max_col, max_row = group.bounds 
        top_left_cell_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(group))   # you need to unmerge before writing (see explanation #1 below)
        for irow in range(min_row, max_row+1):
            for jcol in range(min_col, max_col+1): 
                ws.cell(row = irow, column = jcol, value = top_left_cell_value)
                
    #now "ws" is the excel file unmerged
    #convert openpyxl format to pandas
    data = ws.values
    cols = next(data)[2:]
    data = list(data)
    idx = [r[0:2] for r in data]
    data = (islice(r, 2, None) for r in data)
    multi_index = pd.MultiIndex.from_tuples(idx,
                                           names=['machine','denomination'])
    df = pd.DataFrame(data, index=multi_index, columns=cols)
    
    #keep only machines and operator index
    df = df[~df.index.duplicated(keep='first')]
    index_to_keep = df.index.dropna()
    
    df = df.loc[index_to_keep]
    
    df =df.drop(df.filter(like='None').columns, axis=1)
    df = df.drop(columns = df.columns[0])
    
    #resample the column of df
    columns = pd.to_datetime(df.columns)
    
    resampled_date = pd.DataFrame(0,columns=[""],index=columns).resample('12H', closed = "left").mean().index.tolist()
    df = df.iloc[: , :-1] #drop the last column to make the datetime list fit
    df.columns = resampled_date
    df_t = df.transpose() #transpose to have the date as index
    df_t = df_t.drop(columns = df_t.columns[0:1])#again drop useless columns
    
    
    index_machine = df_t.columns.levels
    prods = []
    for row in df_t.loc[:,(index_machine[0],"N° de dossier")].iterrows():
        for machine in row[1]:
            
            if machine is not None:
                prods.append(machine)
    prods = np.array(prods)
    
    prods = np.unique(prods)


    operator_list = ["SFR", "BPI", "JPI", "JDD", "FDS", "SFH", "NDE", "LTF", "SRG", "JPE", "RPI", "ANA",
                    "MTR", "CMT", "CGR", "CPO", "REA", "LBU"]
    
    merge_occupations = ["OCTODURE", "BP", "TP", "MEL", "EXT",  "BB", "TM", "CF", "MILIEU", 
                         "LYO", "CAPS", "IV", "LAVERIE", "FORMATION", "ETIQUE", 
                         "REQUALIF", "NETTOYAGE", "TEST", "VACANCE", "ABSENT" ,"CONGE",]
    
    to_keep = ["OCTODURE", "BP","TP","EXT","MEL","BB", "TM", "CF", "MILIEU",
                         "LYO", "CAPS","IV", "LAVERIE","AUTRE"]
    
    occupations_operators = {}
    
    for operator in operator_list:
        occupations = []
        for i, row in enumerate(df_t.loc[:,(operator,"Opérateur")]):
            
            if row is not None:
                row = unidecode.unidecode(row)
                row = row.upper()
                for potential in merge_occupations:
                    if potential in row:
                        row = potential
                occupations.append(row)
                
            else:
                if df_t.iloc[i].name.dayofweek<5:
                    #print(df_t.iloc[i].name, op)
                    pass
        
                
        occupations_operators[operator], counts = np.unique(np.array(occupations), return_counts =  True)
        count_sort_ind = np.argsort(-counts)
        occupations_operators[operator] = (occupations_operators[operator][count_sort_ind],counts[count_sort_ind])
        
        
    operation_vs_operator = pd.DataFrame(data = 0, index = operator_list, columns = to_keep + ["ABSENT"])
    
    test = []
    for OPERATOR in operator_list:
    
        for i,o in enumerate(occupations_operators[OPERATOR][0]):
            
            if o in to_keep:
                operation_vs_operator.loc[OPERATOR,o] = operation_vs_operator.loc[OPERATOR,o] + occupations_operators[OPERATOR][1][i]
            elif o in ["VACANCE", "ABSENT","CONGE"]:
                operation_vs_operator.loc[OPERATOR,"ABSENT"] = operation_vs_operator.loc[OPERATOR,"ABSENT"] + occupations_operators[OPERATOR][1][i]
            else:
                operation_vs_operator.loc[OPERATOR,"AUTRE"] = operation_vs_operator.loc[OPERATOR,"AUTRE"] + occupations_operators[OPERATOR][1][i]
                test.append(o)
                
    
    other, counts = np.unique(np.array(test), return_counts =  True)
    count_sort_ind = np.argsort(-counts)
    other = (other[count_sort_ind],counts[count_sort_ind])
    
    nbr_operations = len(operation_vs_operator.columns)
    columns = operation_vs_operator.columns
    
    for column in operation_vs_operator.columns:
        operation_vs_operator[column] = operation_vs_operator[column]  / operation_vs_operator[column].sum()
        
    return operation_vs_operator.T